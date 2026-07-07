"""YouTube 企画・台本 自動生成ツール（杉山先生・借金問題チャンネル）

過去台本(Excel)から抽出した「型」に沿って台本を生成する。
毎回固定の定型文(config/channel.yaml の fixed_blocks)はそのまま挿入し、
テーマ固有部分（掴み・共感・本編・まとめ等）だけを AI(Claude) で生成。
出力は杉山先生の様式の Excel（項目｜内容｜参考記事｜文字数、1行=1ナレーション）。

使い方:
    python src/generate.py ideate --count 6
    python src/generate.py script --theme "任意整理をすると家族にバレるのか"
    python src/generate.py check  --file output/xxx.xlsx
    python src/generate.py run            # 企画→選択→台本(Excel)→点検 一気通し
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path

import yaml
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import anthropic

import usage

# --- 基本設定 ---------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = ROOT / "config" / "channel.yaml"
HISTORY_PATH = ROOT / "config" / "topics_history.txt"
OUTPUT_DIR = ROOT / "output"

MODEL = "claude-opus-4-8"  # 既定モデル（config の models で上書き可能）


def model_for(cfg: dict, task: str) -> str:
    """タスク（ideate/script/review）ごとのモデルを config から取得。"""
    return cfg.get("models", {}).get(task, MODEL)

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

load_dotenv(ROOT / ".env")


def load_config() -> dict:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)


KNOWLEDGE_PATH = ROOT / "config" / "knowledge.yaml"


def load_knowledge() -> list[dict]:
    """杉山先生の実務ナレッジ（テーマ別）。無ければ空。"""
    if not KNOWLEDGE_PATH.exists():
        return []
    with open(KNOWLEDGE_PATH, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    return data.get("knowledge", []) or []


def select_knowledge(theme: str) -> str:
    """テーマに関係する実務ナレッジだけを選び、プロンプト用ブロックにする。"""
    theme = theme or ""
    picked = []
    for entry in load_knowledge():
        kws = entry.get("keywords", []) or []
        if any(kw and kw in theme for kw in kws):
            picked.append(entry)
    if not picked:
        return ""
    lines = [
        "【杉山先生の実務ナレッジ（このテーマに関係する“調べても出にくい”勘所）】",
        "※すべて実務上の“目安”。台本では断定せず「目安として」等を添える。ここに無い数字を捏造しない。",
    ]
    for e in picked:
        lines.append(f"■ {e.get('topic', '')}")
        lines += [f"- {p}" for p in e.get("points", [])]
    return "\n".join(lines) + "\n"


def _read_lines(path: Path) -> list[str]:
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8").splitlines()
    return [ln.strip() for ln in lines if ln.strip() and not ln.startswith("#")]


def load_history() -> list[str]:
    return _read_lines(HISTORY_PATH)


def load_reference() -> list[str]:
    return _read_lines(ROOT / "config" / "reference_videos.txt")


# 直近に生成した企画タイトルを記録し、再読込・再アクセスをまたいで重複を避ける
RECENT_IDEAS_PATH = ROOT / "output" / "recent_ideas.txt"


def load_recent_ideas(n: int = 25) -> list[str]:
    """直近に出した企画タイトル（新しい順で最大n件）。"""
    lines = _read_lines(RECENT_IDEAS_PATH)
    return lines[-n:] if lines else []


def append_recent_ideas(titles: list[str], cap: int = 80) -> None:
    """生成した企画タイトルを追記し、末尾cap件だけ残す。"""
    if not titles:
        return
    existing = _read_lines(RECENT_IDEAS_PATH)
    merged = existing + [t.strip() for t in titles if t and t.strip()]
    merged = merged[-cap:]
    RECENT_IDEAS_PATH.parent.mkdir(exist_ok=True)
    RECENT_IDEAS_PATH.write_text("\n".join(merged) + "\n", encoding="utf-8")


def client() -> anthropic.Anthropic:
    return anthropic.Anthropic()


def first_text(message) -> str:
    for block in message.content:
        if block.type == "text":
            return block.text
    return ""


def parse_json(text: str):
    text = text.strip()
    fence = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if fence:
        text = fence.group(1).strip()
    return json.loads(text)


def extract_json(message):
    """レスポンスの複数テキストブロックから、JSONとして読める最後のものを返す。
    （Web検索を使うと検索の説明文などが混ざるため、後ろから探す）"""
    texts = [b.text for b in message.content if b.type == "text"]
    for t in reversed(texts):
        try:
            return parse_json(t)
        except (json.JSONDecodeError, ValueError):
            continue
    raise ValueError("JSONの取得に失敗しました")


# 最近の話題・トレンドや、官公庁の出典を拾うためのサーバーサイドWebツール
WEB_SEARCH_TOOL = {"type": "web_search_20260209", "name": "web_search"}
WEB_FETCH_TOOL = {"type": "web_fetch_20260209", "name": "web_fetch"}


# ============================================================
# 共通プロンプト部品
# ============================================================
def philosophy_block(cfg: dict) -> str:
    p = cfg["philosophy"]
    core = "\n".join(f"- {c}" for c in p["core"])
    forbidden = "\n".join(f"- {c}" for c in p["forbidden_expressions"])
    return (
        "【杉山先生の価値観（最重要・絶対遵守）】\n"
        f"{core}\n"
        "やってはいけない表現:\n"
        f"{forbidden}\n"
        f"メッセージ性の要件: {p['message_requirement']}\n"
    )


def clickability_block(cfg: dict) -> str:
    cl = cfg["clickability"]
    patterns = "\n".join(f"- {p}" for p in cl["patterns"])
    title_policy = cl.get("title_policy", "")
    return (
        "【伸びる企画の型（テーマ選び・タイトルの基準）】\n"
        f"{cl['principle']}\n"
        f"{patterns}\n"
        + (f"【タイトルの約束を守る（最重要）】\n{title_policy}\n" if title_policy else "")
        + f"注意: {cl['caution']}\n"
    )


def compliance_block(cfg: dict) -> str:
    c = cfg["compliance"]
    principles = "\n".join(f"- {x}" for x in c["principles"])
    return (
        "【コンプライアンス（弁護士法・日弁連広告規程・景表法を意識／厳守）】\n"
        f"{principles}\n"
        f"必須の一言: {c['required_disclaimer']}\n"
    )


def policy_block(cfg: dict) -> str:
    notes = cfg.get("policy_notes", []) or []
    if not notes:
        return ""
    return (
        "\n【自事務所ポリシー（企画・台本ともに厳守。主題化も強調もしない）】\n"
        + "\n".join(f"- {n}" for n in notes)
        + "\n"
    )


def channel_context(cfg: dict) -> str:
    ch = cfg["channel"]
    aud = cfg["audience"]
    segs = "\n".join(f"- {s}" for s in aud["segments"])
    return (
        "あなたは、借金問題を扱う弁護士YouTubeチャンネルの放送作家です。\n"
        f"チャンネル: {ch['name']}\n"
        f"出演: {ch['host']}（一人称は「{ch['first_person']}」）\n"
        f"目的: {ch['purpose']}\n"
        f"主な視聴者: {aud['primary']}\n{segs}\n"
        f"視聴者心理: {aud['mindset']}\n\n"
        + philosophy_block(cfg)
        + policy_block(cfg)
    )


# ============================================================
# 企画出し
# ============================================================
IDEA_SCHEMA = {
    "type": "object",
    "properties": {
        "ideas": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "title": {"type": "string", "description": "クリックされる動画タイトル（視聴者の検索・不安に直結）"},
                    "summary": {"type": "string", "description": "内容サマリ。誰のどんな悩みに、何を答える動画か。1〜2文・80字以内で簡潔に。"},
                    "fresh_angle": {"type": "string", "description": "新しい切り口（最近の話題や独自視点）。40字以内で簡潔に。"},
                },
                "required": ["title", "summary", "fresh_angle"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["ideas"],
    "additionalProperties": False,
}


def ideate(cfg: dict, count: int, use_web: bool = True,
           exclude: list[str] | None = None) -> list[dict]:
    import random
    history = load_history()
    history_block = (
        "【過去に扱ったテーマ（切り口・結論の丸かぶりを避ける）】\n"
        + "\n".join(f"- {t}" for t in history)
        if history else "（過去テーマの記録はまだありません）"
    )
    reference = load_reference()
    # 参照リストはランダムに一部だけ見せる。毎回同じ“鉄板ネタ”に引っ張られるのを防ぐ。
    if len(reference) > 10:
        reference = random.sample(reference, 10)
    reference_block = (
        "【伸びている動画タイトル（これを分析して、共通する“クリックされる型”を抽出する）】\n"
        + "\n".join(f"- {t}" for t in reference)
        if reference else "（参照リストは空です。一般的に伸びる型で考えてください）"
    )
    avoid = cfg.get("avoid_topics", []) or []
    avoid_block = (
        "【絶対に扱わないテーマ（これらを主題にした企画は出さない）】\n"
        + "\n".join(f"- {a['topic']}（理由：{a['reason']}）" for a in avoid)
        + "\n"
        if avoid else ""
    )
    # セッション内の既出（exclude）に、ファイル記録の直近生成分もマージ（再読込をまたいで重複回避）
    exclude = list(dict.fromkeys((exclude or []) + load_recent_ideas()))
    exclude_block = (
        "【すでに出した案（重要）】\n"
        "これらは提示済み。タイトルを言い換えただけの案もNG。"
        "同じテーマ・同じ制度・同じ悩みが主役の案は出さず、まったく別の領域から発想する：\n"
        + "\n".join(f"- {t}" for t in exclude)
        + "\n"
        if exclude else ""
    )
    angles = cfg.get("diversity_angles", []) or []
    system = channel_context(cfg) + "\n" + clickability_block(cfg) + "\n" + compliance_block(cfg)

    def _request(n: int, web: bool) -> list[dict]:
        """1回分の企画生成。呼ぶたびに切り口・多様性シードを変える。"""
        web_instruction = (
            "0. まずWeb検索で、最近SNS（X等）やニュースで話題になっている借金・債務整理・お金まわりの\n"
            "   トピック（例：「破産者マップ」など）を調べ、企画の新しい切り口として活かす。\n"
            if web else ""
        )
        angle_block = ""
        if angles:
            picked = random.sample(angles, min(4, len(angles)))
            angle_block = (
                "【今回は特にこの領域から優先して発想を広げる（毎回変わる。ここに寄せて案を作る）】\n"
                + "\n".join(f"- {a}" for a in picked) + "\n"
            )
        user = (
            f"このチャンネルの次回以降の企画案を{n}本出してください（多めに出す。互いにテーマを散らす）。\n\n"
            "手順:\n"
            f"{web_instruction}"
            "1. 下の『伸びている動画タイトル』を分析し、共通する“クリックされる型”を見抜く。\n"
            "2. その型を使いつつ、過去の焼き直しではなく『新しい切り口』を加えた企画を作る。\n"
            "   （最近の話題・世の中の動き・新しい制度や事件に接続できると強い）\n\n"
            "条件:\n"
            "- 先生が話したいことより、視聴者が見たい・検索することを優先する（需要ドリブン）。\n"
            "- タイトルは具体的な不安・疑問・損得に直結させる。歴史の解説に寄せすぎない。\n"
            "- ただし債務者を貶める煽り（「借金の末路」等）は使わない。怖がらせて終わらせず、最後は安心と次の一歩へ。\n"
            "- 本編で視聴者が持ち帰れる価値（行動・損得・判断基準など）を必ず用意する。\n"
            "- 各案は必ず実在しうる具体的な企画にする。『ダミー』『サンプル』などの穴埋め案は絶対に入れない。\n"
            "- 過去テーマと切り口・結論が丸かぶりしないこと。各案の fresh_angle に“何が新しいか”を必ず書く。\n\n"
            f"{avoid_block}\n"
            f"{angle_block}\n"
            f"{exclude_block}\n"
            f"{reference_block}\n\n"
            f"{history_block}\n\n"
            f"（多様性シード:{random.randint(1000, 9999)}。このシードは毎回変わる。"
            f"前回と同じ発想に固執せず、上の『今回優先して広げる領域』に寄せて、"
            f"毎回できるだけ違う案を出すこと。案どうしもテーマを散らす。）"
        )
        params = dict(
            model=model_for(cfg, "ideate"),
            max_tokens=8000,  # 検索＋複数案でも出力が切れないよう十分に確保
            system=system,
            temperature=1.0,
            thinking={"type": "disabled"},  # 企画出しは速度優先（深い思考は不要）
            output_config={"format": {"type": "json_schema", "schema": IDEA_SCHEMA}},
            messages=[{"role": "user", "content": user}],
        )
        if web:
            params["tools"] = [WEB_SEARCH_TOOL]
        msg = client().messages.create(**params)
        usage.record(cfg, params["model"], msg.usage, "ideate")
        return extract_json(msg).get("ideas", []) or []

    def _is_junk(title: str) -> bool:
        t = (title or "").strip()
        return len(t) < 8 or t.lower() in {
            "ダミー", "dummy", "サンプル", "sample", "テスト", "未定", "タイトル", "例"}

    # 生成→ジャンク/重複除去。count件に満たなければ1回だけ補充（2回目は速度優先で検索なし）。
    seen_keys = {t.strip().lower() for t in exclude}
    collected: list[dict] = []
    for attempt in range(2):
        for d in _request(count + 2, use_web and attempt == 0):
            key = str(d.get("title", "")).strip().lower()
            if not key or key in seen_keys or _is_junk(d.get("title", "")):
                continue
            seen_keys.add(key)
            collected.append(d)
        if len(collected) >= count:
            break
    result = collected[:count]
    append_recent_ideas([d.get("title", "") for d in result])  # 次回以降の重複回避に記録
    return result


def print_ideas(ideas: list[dict]) -> None:
    for i, idea in enumerate(ideas, 1):
        print(f"\n━━━ 企画 {i} ━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"タイトル     : {idea['title']}")
        print(f"内容サマリ   : {idea.get('summary', '')}")
        print(f"新しい切り口 : {idea.get('fresh_angle', '')}")


# ============================================================
# 台本生成（テーマ固有部分のみ生成 → 定型と合成）
# ============================================================
# 本編の1まとまり（ブロック）：本文の行 + その内容を裏づける出典
BODY_BLOCK = {
    "type": "object",
    "properties": {
        "lines": {"type": "array", "items": {"type": "string"},
                  "description": "1行=1ナレーション(15〜50字)。話題が一つのまとまり。"},
        "references": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "資料名・ページ名（例：金融庁『多重債務者向け相談窓口』）"},
                    "url": {"type": "string", "description": "Web検索で実際に得たURL。実在しないURLは絶対に書かない。"},
                },
                "required": ["name", "url"],
                "additionalProperties": False,
            },
            "description": "このまとまりの事実を裏づける出典。Web検索で実在を確認したものだけ。官公庁・公的機関を最優先。無ければ空配列でよい（捏造厳禁）。",
        },
    },
    "required": ["lines", "references"],
    "additionalProperties": False,
}

SCRIPT_SCHEMA = {
    "type": "object",
    "properties": {
        "title": {"type": "string", "description": "クリックされる動画タイトル"},
        "thumbnail": {"type": "string", "description": "サムネに載せる短いコピー"},
        "theme_label": {"type": "string", "description": "テーマ明言の「」内に入れる短い表現"},
        "hook": {"type": "string", "description": "最初の引き（冒頭1行のフック）"},
        "empathy": {"type": "array", "items": {"type": "string"},
                    "description": "視聴者の気持ちを代弁する行（5〜10行）"},
        "gap": {"type": "array", "items": {"type": "string"},
                "description": "認識のギャップを突いて視聴意欲を掻き立てる行（3〜7行）"},
        "expert_bridge": {"type": "string",
                          "description": "本編の解説宣言につなぐ『前フリ』の1行。"
                          "直後に『現役の弁護士である私杉山が、いつものように解りやすく解説していきたいと思います。』という決まり文句が自動で入る。"
                          "そのため、この行には『私杉山が』『解説します／解説していきます』を絶対に含めない（重複するため）。"
                          "『今回は、〜について、日頃から〜を数多く扱っている』のように連用形（〜ている）で止めて、次の決まり文句へ自然につなぐこと。"},
        "viewer_types": {"type": "array", "items": {"type": "string"},
                         "description": "視聴人物像（こういう方は最後まで見て、の対象）3パターン分の行"},
        "body_part1": {"type": "array", "items": BODY_BLOCK,
                       "description": "本編の前半。話題のまとまり(ブロック)の配列。各ブロックは本文の行と出典を持つ。"
                                      "最後のブロックは話題が一区切りつく自然な切れ目にする（直後に中盤LINE誘導が入る）。"},
        "body_part2": {"type": "array", "items": BODY_BLOCK,
                       "description": "本編の後半（中盤LINE誘導のあと）。前半とは別の論点・次の話題から始める。同じくブロックの配列。"},
        "summary": {"type": "array", "items": {"type": "string"},
                    "description": "まとめ本文。本編の要点を振り返る行（6〜12行）"},
    },
    "required": ["title", "thumbnail", "theme_label", "hook", "empathy", "gap",
                 "expert_bridge", "viewer_types", "body_part1", "body_part2", "summary"],
    "additionalProperties": False,
}


def script_system(cfg: dict) -> str:
    spec = cfg["script_spec"]
    voice = cfg["voice"]
    style = "\n".join(f"- {s}" for s in voice["style"])
    guide = cfg.get("style_guide", []) or []
    guide_block = (
        "\n【杉山先生の型（過去台本191本＋作成マニュアルから抽出。できる範囲で従う）】\n"
        + "\n".join(f"- {g}" for g in guide)
        + "\n"
        if guide else ""
    )
    return (
        channel_context(cfg)
        + "\n【語り口】\n"
        + f"トーン: {voice['tone']}\n{style}\n"
        + guide_block
        + "\n"
        + "【台本の作り方】\n"
        + f"- 1行=1ナレーション（声に出して自然な一息）。1行は{spec['line_max_chars']}字以内を目安に短く区切る。\n"
        + f"- これは杉山先生が加筆する『下書き』。無理に文字数を埋めない。要点は深く書くが、先生が書き足す余地は残してよい（全体で{spec['recommend_total_chars']}字前後を上限の目安）。\n"
        + "- 専門用語は必ず直後にやさしく言い換える。\n"
        + "- 制度の説明だけで終わらせず、視聴者が持ち帰れる価値（具体的な行動・損得・不安の解消・判断の基準）を必ず入れる。\n"
        + "\n【深さと具体性（杉山先生の台本に近づけるための肝）】\n"
        + "- タイトルの約束を果たす：タイトルが視聴者に約束している核心の問い・期待に、本編で正面から深く答える。タイトルは『家を守る方法』なのに本編が別制度の一般紹介、のようなズレを絶対に作らない。\n"
        + "- 一点集中で深掘り：関連する制度を広く浅く並べる『制度カタログ』にしない。今回のタイトルの核心を1つに絞り、その1点を具体例・数字・条文・『なぜ』で厚く掘る。他の制度は核心を引き立てる範囲で最小限に触れる。\n"
        + "- なぜを必ず添える：制度・ルール・結論を述べたら、必ず『なぜそうなっているのか（立法趣旨・目的・背景）』をワンセットで説明する。事実や結論の列挙で終わらせない。\n"
        + "- 具体性：年は西暦（元号）を併記する。制度の変遷は時系列で『いつ・何が』を具体的に。条文番号・具体的な数値・固有名詞を可能な限り入れ、抽象語で流さない。\n"
        + "- 事件は物語に：テーマに関わる象徴的な事件・実例が1つあれば、物語として深掘りする（経緯→当事者の言い分→それへの反論→社会の対応→そこから得られる普遍的な教訓）。\n"
        + "- 一段深い学び：動画のどこかで、視聴者が『なるほど』と思える一段抽象度の高い原則・視点を1つ提示する（単なる制度紹介を超える）。\n"
        + "- 行間を繋ぐ：理由・背景は『〜のですが』『〜わけです』『〜からです』で前後を論理的につなぐ。短い断定文の羅列にしない（1行を短く、は維持）。\n"
        + "- 実務トピックを1つ：信用情報（ブラックリスト）など視聴者に身近な実務トピックを1つ添えてよい。ただし主題を薄めない範囲にとどめ、主題の深掘りを優先する。\n\n"
        + "- 歴史や社会的背景は、テーマの核心が制度の変遷・事件にある場合は時系列で具体的に掘る。そこに核心がないテーマでは無理に足さない。\n"
        + "- 怖いテーマでも、最後は視聴者を安心させ、次の一歩を示す。債務者を貶めない。\n"
        + "- 本編は body_part1 と body_part2 に分ける。前半は話題が一区切りつく自然な切れ目で終える（その直後に中盤LINE誘導が入る）。後半は次の論点から始める。文の途中で割らない。\n"
        + "- どこかで必須の免責の一言を自然に入れる。\n"
        + "- 定型文（自己紹介・LINE誘導・まとめCTA・ED）はシステム側で自動挿入するので、生成しないこと。\n\n"
        + compliance_block(cfg)
    )


def generate_parts(cfg: dict, theme: str, use_web: bool = True) -> dict:
    history = load_history()
    history_block = (
        "【過去に扱ったテーマ（切り口・結論の丸かぶりを避ける）】\n"
        + "\n".join(f"- {t}" for t in history)
        if history else "（過去テーマの記録はまだありません）"
    )
    if use_web:
        cite_instruction = (
            "【出典（信頼性のため重要）】\n"
            "- 事実・数字・制度・期間などを述べる本編のまとまりでは、Web検索で出典を調べ、各ブロックの references に入れる。\n"
            "- 官公庁・公的機関を最優先（例：法務省、裁判所、金融庁、消費者庁、国民生活センター、日本弁護士連合会、e-Gov法令検索、官報、e-Stat など）。\n"
            "- 速度優先のため、ページ全文の精読はせず、検索結果に出た実在の公的機関の資料名・URLを引用する。\n"
            "- 実在を確認したURLだけを書く。URL・資料名の捏造は絶対にしない。確証がなければ references は空配列にする。\n"
            "- 冒頭・LINE誘導・まとめ・ED には出典は不要（references は空配列）。\n"
        )
    else:
        cite_instruction = (
            "【出典】今回はネット検索を使わないため、references はすべて空配列にする（URLを推測で書かない）。\n"
        )
    knowledge_block = select_knowledge(theme)
    knowledge_instruction = (
        f"{knowledge_block}\n"
        "↑この実務ナレッジを本編に自然に織り込み、"
        "『ネットで調べれば出る一般論』を超える具体（相場・目安・落とし穴・実務判断）を必ず盛り込むこと。"
        "ただし目安は断定せず、ここに無い数字は作らない。\n\n"
        if knowledge_block else ""
    )
    user = (
        f"次のテーマで、台本のテーマ固有部分を作ってください。\n"
        f"テーマ: 「{theme}」\n\n"
        "制度説明だけに終わらせず、視聴者が持ち帰れる価値（具体的な行動・損得・不安の解消・判断の基準）を必ず入れてください。\n"
        "怖いテーマでも最後は安心と次の一歩へ。債務者を絶対に貶めないでください。\n\n"
        f"{knowledge_instruction}"
        f"{cite_instruction}\n"
        f"{history_block}"
    )
    params = dict(
        model=model_for(cfg, "script"),
        max_tokens=20000,
        system=script_system(cfg),
        thinking={"type": "adaptive"},
        output_config={"format": {"type": "json_schema", "schema": SCRIPT_SCHEMA}},
        messages=[{"role": "user", "content": user}],
    )
    if use_web:
        params["tools"] = [WEB_SEARCH_TOOL]  # 速度優先：ページ精読(web_fetch)はしない
    with client().messages.stream(**params) as stream:
        for _ in stream.text_stream:
            print(".", end="", flush=True)  # 進捗表示
        msg = stream.get_final_message()
    print()
    usage.record(cfg, params["model"], msg.usage, "script")
    return extract_json(msg)


# --- 行の組み立て ----------------------------------------------
class Row:
    def __init__(self, item: str, text: str, green: bool = False,
                 spacer: bool = False, reference: str = ""):
        self.item = item            # 項目（役割ラベル）。空文字なら継続行
        self.text = text            # 内容（ナレーション）
        self.green = green          # LINE誘導＝緑背景
        self.spacer = spacer        # 区切りの空行
        self.reference = reference  # 参考記事（出典の資料名＋URL）


def assemble_rows(cfg: dict, theme: str, parts: dict) -> list[Row]:
    fb = cfg["fixed_blocks"]
    sf = cfg["semi_fixed"]
    label = parts.get("theme_label") or theme
    rows: list[Row] = []

    def add(item, text, green=False):
        rows.append(Row(item, text, green=green))

    def add_block(item, lines, green=False):
        for i, ln in enumerate(lines):
            add(item if i == 0 else "", ln, green=green)

    def add_body(item, blocks):
        """本編ブロック（本文＋出典）を行に展開。出典は各ブロック先頭行の参考記事欄へ。"""
        first_row = True
        for block in blocks:
            refs = block.get("references", []) or []
            ref_text = "\n".join(f"{r['name']}\n{r['url']}" for r in refs)
            for li, ln in enumerate(block.get("lines", [])):
                rows.append(Row(
                    item if first_row else "",
                    ln,
                    reference=ref_text if li == 0 else "",
                ))
                first_row = False

    def spacer():
        rows.append(Row("", "", spacer=True))

    # 1. 最初の引き＆自己紹介
    add("最初の引き＆自己紹介", parts["hook"])
    for ln in fb["self_intro"]:
        add("", ln)
    # 2. テーマ明言
    add("テーマ明言", sf["theme_statement"].format(theme=label))
    # 3. 視聴者の気持ちを代弁
    add_block("視聴者の気持ちを代弁", parts["empathy"])
    # 4. 認識のギャップ
    add_block("認識のギャップを突いて視聴意欲掻き立て", parts["gap"])
    add("", parts["expert_bridge"])
    add("", sf["expert_declaration"])
    # 5. 視聴人物像の明確化
    add_block("視聴人物像の明確化", parts["viewer_types"])
    add("", sf["viewer_close"])
    spacer()
    # 6. （LINE誘導）冒頭
    add_block("（LINE誘導）", fb["cta_intro"], green=True)
    spacer()
    # 7. 本編前半
    add_body("本編", parts["body_part1"])
    spacer()
    # 8. 中盤 LINE誘導
    add_block("LINE誘導", fb["cta_mid"], green=True)
    spacer()
    # 9. 本編後半
    add_body("本編（続き）", parts["body_part2"])
    spacer()
    add("", sf["transition_to_summary"])
    # 10. まとめ（CTA定型 → まとめ本文）
    add_block("まとめ", fb["closing_cta"], green=True)
    spacer()
    add("", sf["summary_open"].format(theme=label))
    for ln in parts["summary"]:
        add("", ln)
    spacer()
    # 11. ED
    add_block("ED", fb["ending"])
    return rows


# --- Excel 出力 ------------------------------------------------
def write_excel(path: Path, cfg: dict, parts: dict, rows: list[Row]) -> None:
    spec = cfg["script_spec"]
    total = sum(len(r.text) for r in rows if not r.spacer)
    minutes = round(total / spec["chars_per_minute"], 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テンプレート"

    # --- ヘッダ・メタ情報 ---
    ws["A1"] = "タイトル名（仮）"
    ws["B1"] = "●" + parts["title"]
    ws["E1"] = "総文字数"
    ws["F1"] = total
    ws["A2"] = "作成者"
    ws["B2"] = "AI下書き（要・杉山先生確認）"
    ws["A3"] = "想定サムネ"
    ws["B3"] = parts.get("thumbnail", "")
    ws["A4"] = "想定ターゲット"
    ws["B4"] = cfg["audience"]["primary"]
    ws["E4"] = "想定時間 ※450文字≒1分。4,500文字≒10分を超えるように"
    ws["F4"] = f"{minutes}分"
    ws["A5"] = "※表表示の際は行の背景を黄色に"
    ws["A6"] = "※LINE表示の際は背景を緑色に"
    ws["A7"] = "※参考記事は2記事以上ご使用ください。（同じ記事の同じ内容・構成順番にしない）"
    ws["A8"] = "▼参考記事を必ず記載する（冒頭とED以外は必ず参考元をご記載ください。）"

    # --- 表ヘッダ ---
    head_row = 9
    headers = ["NO", "NO2", "項目", "内容", "参考記事", "文字数"]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=head_row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)

    # --- 本文行 ---
    r = head_row + 1
    no = 2
    for row in rows:
        ws.cell(row=r, column=1, value=no)  # NO
        if not row.spacer:
            if row.item:
                ws.cell(row=r, column=3, value=row.item)  # 項目
            c_naiyou = ws.cell(row=r, column=4, value=row.text)  # 内容
            c_naiyou.alignment = Alignment(wrap_text=True, vertical="top")
            if row.reference:
                c_ref = ws.cell(row=r, column=5, value=row.reference)  # 参考記事（出典）
                c_ref.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=6, value=len(row.text))  # 文字数
            if row.green:
                for col in range(1, 7):
                    ws.cell(row=r, column=col).fill = GREEN
        else:
            ws.cell(row=r, column=6, value=0)
        r += 1
        no += 1

    # 体裁
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 18

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(path)


def assembled_text(rows: list[Row]) -> str:
    return "\n".join(r.text for r in rows if not r.spacer and r.text)


def output_path(title: str) -> Path:
    today = dt.date.today().isoformat()
    slug = re.sub(r"[^\w一-龠ぁ-んァ-ンー]+", "_", title)[:30].strip("_")
    return OUTPUT_DIR / f"{today}_{slug}_AI下書き.xlsx"


# ============================================================
# コンプラ点検
# ============================================================
ISSUE_SCHEMA = {
    "type": "object",
    "properties": {
        "issues": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "quote": {"type": "string"},
                    "problem": {"type": "string"},
                    "law": {"type": "string", "description": "観点（弁護士法/景表法/債務者を貶める/不安煽り 等）"},
                    "suggestion": {"type": "string"},
                    "severity": {"type": "string", "enum": ["高", "中", "低"]},
                },
                "required": ["quote", "problem", "law", "suggestion", "severity"],
                "additionalProperties": False,
            },
        },
        "overall": {"type": "string"},
    },
    "required": ["issues", "overall"],
    "additionalProperties": False,
}


def local_ng_scan(cfg: dict, body: str) -> list[str]:
    return [p for p in cfg["compliance"]["ng_phrases"] if p in body]


def llm_review(cfg: dict, body: str) -> dict:
    system = channel_context(cfg) + "\n" + compliance_block(cfg)
    user = (
        "あなたは弁護士広告に詳しい校閲者です。次の台本を、弁護士法・日弁連の広告規程・"
        "景品表示法・過度な不安喚起・債務者を貶める表現の観点で点検し、問題箇所を指摘してください。"
        "問題がなければ issues を空配列にしてください。\n\n"
        "----- 台本ここから -----\n"
        f"{body}\n"
        "----- 台本ここまで -----"
    )
    msg = client().messages.create(
        model=model_for(cfg, "review"),
        max_tokens=8000,
        system=system,
        thinking={"type": "disabled"},  # チェック用途。思考でトークンを使い切って空応答になるのを防ぐ
        output_config={"format": {"type": "json_schema", "schema": ISSUE_SCHEMA}},
        messages=[{"role": "user", "content": user}],
    )
    usage.record(cfg, model_for(cfg, "review"), msg.usage, "review")
    # レビューが空・解析不能でも、台本本体は必ず届ける（チェック未実行として返す）
    try:
        return extract_json(msg)
    except (ValueError, json.JSONDecodeError):
        return {"issues": [], "overall": "※自動チェックを実行できませんでした。公開前に内容を必ずご確認ください。"}


def run_check(cfg: dict, body: str) -> None:
    print("\n===== コンプラ点検 =====")
    ng = local_ng_scan(cfg, body)
    if ng:
        print("⚠ NGワード検知（要修正）:", "、".join(ng))
    else:
        print("✓ NGワードの機械検知: なし")
    review = llm_review(cfg, body)
    issues = review["issues"]
    if not issues:
        print("✓ AIレビュー: 重大な問題は見つかりませんでした")
    else:
        print(f"⚠ AIレビュー: {len(issues)} 件の指摘")
        for n, it in enumerate(issues, 1):
            print(f"\n  [{n}] 重要度{it['severity']}／{it['law']}")
            print(f"      原文 : {it['quote']}")
            print(f"      問題 : {it['problem']}")
            print(f"      修正案: {it['suggestion']}")
    print(f"\n総評: {review['overall']}")


def read_xlsx_text(path: Path) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=4, max_col=4, values_only=True):
            if row[0] and str(row[0]) != "内容":
                out.append(str(row[0]))
    return "\n".join(out)


# ============================================================
# サブコマンド
# ============================================================
def make_script(cfg: dict, theme: str, use_web: bool = True) -> tuple[Path, str]:
    print(f"\n台本を生成します: 「{theme}」"
          + ("（出典をネット検索しながら作成）" if use_web else ""))
    print("（生成中", end="", flush=True)
    parts = generate_parts(cfg, theme, use_web=use_web)
    print("生成完了）")
    rows = assemble_rows(cfg, theme, parts)
    path = output_path(parts["title"])
    write_excel(path, cfg, parts, rows)
    total = sum(len(r.text) for r in rows if not r.spacer)
    print(f"\nタイトル: {parts['title']}")
    print(f"サムネ  : {parts.get('thumbnail', '')}")
    print(f"総文字数: {total}（約 {round(total / cfg['script_spec']['chars_per_minute'], 1)} 分）")
    print(f"保存先  : {path}")
    return path, assembled_text(rows)


def cmd_ideate(args) -> None:
    print_ideas(ideate(load_config(), args.count, use_web=not args.no_web))


def cmd_script(args) -> None:
    cfg = load_config()
    _, body = make_script(cfg, args.theme, use_web=not args.no_web)
    if not args.no_check:
        run_check(cfg, body)


def cmd_check(args) -> None:
    cfg = load_config()
    p = Path(args.file)
    body = read_xlsx_text(p) if p.suffix == ".xlsx" else p.read_text(encoding="utf-8")
    run_check(cfg, body)


def cmd_run(args) -> None:
    cfg = load_config()
    print("企画案を出します...")
    ideas = ideate(cfg, args.count, use_web=not args.no_web)
    print_ideas(ideas)
    choice = input(f"\n台本にする企画の番号を選んでください (1-{len(ideas)}): ").strip()
    try:
        idea = ideas[int(choice) - 1]
    except (ValueError, IndexError):
        print("番号が不正です。終了します。")
        return
    _, body = make_script(cfg, idea["title"], use_web=not args.no_web)
    run_check(cfg, body)


def main() -> None:
    parser = argparse.ArgumentParser(description="YouTube 企画・台本 自動生成ツール（杉山先生チャンネル）")
    sub = parser.add_subparsers(dest="command", required=True)

    p_ideate = sub.add_parser("ideate", help="企画案を出す")
    p_ideate.add_argument("--count", type=int, default=6)
    p_ideate.add_argument("--no-web", action="store_true", help="ネット検索（トレンド反映）を使わない")
    p_ideate.set_defaults(func=cmd_ideate)

    p_script = sub.add_parser("script", help="テーマから台本(Excel)を生成する")
    p_script.add_argument("--theme", required=True, help="台本にするテーマ")
    p_script.add_argument("--no-check", action="store_true", help="コンプラ点検をスキップ")
    p_script.add_argument("--no-web", action="store_true", help="出典のネット検索を使わない")
    p_script.set_defaults(func=cmd_script)

    p_check = sub.add_parser("check", help="既存の台本(xlsx/txt)をコンプラ点検")
    p_check.add_argument("--file", required=True)
    p_check.set_defaults(func=cmd_check)

    p_run = sub.add_parser("run", help="企画→選択→台本(Excel)→点検 一気通し")
    p_run.add_argument("--count", type=int, default=6)
    p_run.add_argument("--no-web", action="store_true", help="ネット検索（トレンド反映）を使わない")
    p_run.set_defaults(func=cmd_run)

    args = parser.parse_args()
    try:
        args.func(args)
    except anthropic.AuthenticationError:
        print("APIキーが不正です。.env の ANTHROPIC_API_KEY を確認してください。", file=sys.stderr)
        sys.exit(1)
    except anthropic.RateLimitError:
        print("レート制限です。少し待って再実行してください。", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
